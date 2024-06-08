import openpyxl


book = openpyxl.load_workbook("D:\\selenium_python\\data\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
# writing to Excel sheet
sheet.cell(row=2, column=2).value = "Ankit"
#
print(sheet.cell(row=2, column=2).value)
# to check how many rows in sheet level (not cell level)
print(sheet.max_row)
# to check how many column in sheet level (not cell level)
print(sheet.max_column)
# To check by cell name
print(sheet['A5'].value)
print(sheet['B2'].value)
print(sheet['C2'].value)
print(sheet['B3'].value)
#
for i in range(1, sheet.max_row + 1):  # to get rows
    if sheet.cell(row=i, column=1).value == "Testcase2":

        for j in range(2, sheet.max_column + 1):  # to get columns
            # Dict["lastname"]="Shukla"
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
